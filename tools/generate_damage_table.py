# -*- coding: utf-8 -*-
"""
生成战斗伤害计算 Excel 表格（完整设计版：含属性伤害+属性抗性+通用抗性）

列结构（38列）：
=== 第一部分：人物属性（16列，A-P）===
A 攻击力  B 攻击速度  C 最大生命  D 暴击率(%)  E 暴击伤害(%)  F 防御  G 物理穿透
H 火属性伤害  I 冰属性伤害  J 雷属性伤害  K 毒属性伤害
L 通用抗性(%)  M 火抗性(%)  N 冰抗性(%)  O 雷抗性(%)  P 毒抗性(%)

=== 第二部分：怪物属性（14列，Q-AD）===
Q 攻击力  R 最大生命  S 防御  T 暴击率(%)  U 暴击伤害(%)  V 攻击速度
W 火属性攻击  X 冰属性攻击  Y 雷属性攻击  Z 毒属性攻击
AA 火抗性(%)  AB 冰抗性(%)  AC 雷抗性(%)  AD 毒抗性(%)

=== 第三部分：伤害计算（8列，AE-AL）===
AE 人物攻击怪物 非暴击总伤害
AF 人物攻击怪物 暴击总伤害
AG 人物击杀怪物 次数(向上取整)
AH 人物击杀怪物 时间(秒)
AI 怪物攻击人物 非暴击总伤害
AJ 怪物攻击人物 暴击总伤害
AK 怪物击杀人物 次数(向上取整)
AL 怪物击杀人物 时间(秒)

核心公式（完整设计版）：
  人物攻击怪物：
    物理基础 = 攻击力 × (1 - max(0,怪防-穿透)/100)
    元素实际 = Σ 每种元素伤害 × (1 - 对应怪物抗性/100)
    非暴击总伤害 = max(1, round(物理基础 + 元素实际, 0))
    暴击总伤害   = max(1, round((物理基础 + 元素实际) × (1+暴伤/100), 0))
    击杀次数 = ceil(怪命 / 非暴击总伤害)
    击杀时间 = 次数 × (1000/攻速) / 1000

  怪物攻击人物：
    通用减伤系数 = 1 / (1 + 通用抗性/100)   （抗性100=减伤50%，抗性200=减伤66.7%，曲线平滑无免伤上限）
    物理基础 = 怪攻 × (1 - 玩家防御/100) × 通用减伤系数
    元素实际 = Σ 每种怪元素攻击 × (1 - 玩家对应抗性/100) × 通用减伤系数
    非暴击总伤害 = max(1, round(物理基础 + 元素实际, 0))
    暴击总伤害   = max(1, round((物理基础 + 元素实际) × (1+怪暴伤/100), 0))
    击杀次数 = ceil(玩家命 / 非暴击总伤害)
    击杀时间 = 次数 × (1000/怪攻速) / 1000
"""
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "伤害计算表"

# ===== 三部分淡色填充 =====
fill_player   = PatternFill("solid", fgColor="E8F0FE")  # 淡蓝
fill_monster  = PatternFill("solid", fgColor="FCE8E6")  # 淡红
fill_calc     = PatternFill("solid", fgColor="E6F4EA")  # 淡绿
fill_header   = PatternFill("solid", fgColor="3C4043")  # 深灰大标题
fill_subhead  = PatternFill("solid", fgColor="80868B")  # 中灰列标题

# 标题内再分"物理/元素/抗性"三小组
fill_sub_p_phys = PatternFill("solid", fgColor="C9DAF8")   # 人物内"物理属性"稍深蓝
fill_sub_p_elem = PatternFill("solid", fgColor="D9EAD3")   # 人物内"元素伤害"稍深绿
fill_sub_p_res  = PatternFill("solid", fgColor="F4ECDC")   # 人物内"抗性"米色
fill_sub_m_phys = PatternFill("solid", fgColor="F4CCCC")   # 怪物内"物理"稍深红
fill_sub_m_elem = PatternFill("solid", fgColor="D9EAD3")   # 怪物内"元素攻击"稍深绿
fill_sub_m_res  = PatternFill("solid", fgColor="F4ECDC")   # 怪物内"抗性"米色

header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
subhead_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
group_font   = Font(name="微软雅黑", size=10, bold=True, color="3C4043")
data_font = Font(name="微软雅黑", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===== 三部分列边界 =====
# 第一部分：16列 A-P；再分 物理7(A-G) + 元素4(H-K) + 抗性5(L-P)
P_PHYS_END, P_ELEM_END, P_RES_END, P_TOTAL = 7, 11, 16, 16
# 第二部分：14列 Q-AD；再分 物理6(Q-V) + 元素4(W-Z) + 抗性4(AA-AD)
M_PHYS_END, M_ELEM_END, M_RES_END, M_TOTAL = P_TOTAL+6, P_TOTAL+10, P_TOTAL+14, P_TOTAL+14
# 第三部分：8列 AE-AL
C_TOTAL = M_TOTAL + 8
TOTAL_COLS = C_TOTAL

# ===== 第1行：三大分组标题 =====
ws.row_dimensions[1].height = 26
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=P_TOTAL)
c = ws.cell(row=1, column=1, value="一、人物属性")
c.font = header_font; c.alignment = center; c.fill = fill_header

ws.merge_cells(start_row=1, start_column=P_TOTAL+1, end_row=1, end_column=M_TOTAL)
c = ws.cell(row=1, column=P_TOTAL+1, value="二、怪物属性")
c.font = header_font; c.alignment = center; c.fill = fill_header

ws.merge_cells(start_row=1, start_column=M_TOTAL+1, end_row=1, end_column=C_TOTAL)
c = ws.cell(row=1, column=M_TOTAL+1, value="三、伤害计算（完整设计版·自动生成）")
c.font = header_font; c.alignment = center; c.fill = fill_header

# ===== 第2行：人物/怪物 内部分组小标题 =====
ws.row_dimensions[2].height = 22
# 人物 物理7
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=P_PHYS_END)
c = ws.cell(row=2, column=1, value="基础属性")
c.font = group_font; c.alignment = center; c.fill = fill_sub_p_phys; c.border = border
# 人物 元素4
ws.merge_cells(start_row=2, start_column=P_PHYS_END+1, end_row=2, end_column=P_ELEM_END)
c = ws.cell(row=2, column=P_PHYS_END+1, value="属性伤害（人物→怪物）")
c.font = group_font; c.alignment = center; c.fill = fill_sub_p_elem; c.border = border
# 人物 抗性5
ws.merge_cells(start_row=2, start_column=P_ELEM_END+1, end_row=2, end_column=P_RES_END)
c = ws.cell(row=2, column=P_ELEM_END+1, value="抗性（减伤怪物伤害）")
c.font = group_font; c.alignment = center; c.fill = fill_sub_p_res; c.border = border

# 怪物 物理6
ws.merge_cells(start_row=2, start_column=P_TOTAL+1, end_row=2, end_column=M_PHYS_END)
c = ws.cell(row=2, column=P_TOTAL+1, value="基础属性")
c.font = group_font; c.alignment = center; c.fill = fill_sub_m_phys; c.border = border
# 怪物 元素4
ws.merge_cells(start_row=2, start_column=M_PHYS_END+1, end_row=2, end_column=M_ELEM_END)
c = ws.cell(row=2, column=M_PHYS_END+1, value="属性攻击（怪物→人物）")
c.font = group_font; c.alignment = center; c.fill = fill_sub_m_elem; c.border = border
# 怪物 抗性4
ws.merge_cells(start_row=2, start_column=M_ELEM_END+1, end_row=2, end_column=M_RES_END)
c = ws.cell(row=2, column=M_ELEM_END+1, value="抗性（减伤人物属性伤害）")
c.font = group_font; c.alignment = center; c.fill = fill_sub_m_res; c.border = border

# 第三部分第2行合并
ws.merge_cells(start_row=2, start_column=M_TOTAL+1, end_row=2, end_column=C_TOTAL)
c = ws.cell(row=2, column=M_TOTAL+1, value="伤害/击杀次数/击杀时间（含属性+抗性）")
c.font = group_font; c.alignment = center; c.fill = fill_calc; c.border = border

# ===== 第3行：各列标题 =====
ws.row_dimensions[3].height = 46
player_headers = [
    "攻击力","攻击速度","最大生命","暴击率\n(%)","暴击伤害\n(%)","防御","物理穿透",
    "火属性\n伤害","冰属性\n伤害","雷属性\n伤害","毒属性\n伤害",
    "通用抗性\n(%)","火抗性\n(%)","冰抗性\n(%)","雷抗性\n(%)","毒抗性\n(%)",
]
monster_headers = [
    "攻击力","最大生命","防御","暴击率\n(%)","暴击伤害\n(%)","攻击速度",
    "火属性\n攻击","冰属性\n攻击","雷属性\n攻击","毒属性\n攻击",
    "火抗性\n(%)","冰抗性\n(%)","雷抗性\n(%)","毒抗性\n(%)",
]
calc_headers = [
    "人物攻击怪物\n非暴击总伤害","人物攻击怪物\n暴击总伤害","人物击杀怪物\n次数(ceil)","人物击杀怪物\n时间(秒)",
    "怪物攻击人物\n非暴击总伤害","怪物攻击人物\n暴击总伤害","怪物击杀人物\n次数(ceil)","怪物击杀人物\n时间(秒)",
]

all_headers = player_headers + monster_headers + calc_headers
for idx, name in enumerate(all_headers, start=1):
    cell = ws.cell(row=3, column=idx, value=name)
    cell.font = subhead_font; cell.alignment = center; cell.border = border
    cell.fill = fill_subhead  # 基础深灰，下面按组染色
# 按部分覆盖列标题背景色
for col in range(1, P_TOTAL+1):
    ws.cell(row=3, column=col).fill = fill_subhead if col <= P_PHYS_END else (fill_subhead if col<=P_ELEM_END else fill_subhead)
# 边框都已有，保持统一深灰标题即可

# ===== 5条测试数据 =====
# [人攻,攻速,生命,暴率,暴伤,防,穿透, 火伤,冰伤,雷伤,毒伤, 通抗,火抗,冰抗,雷抗,毒抗,
#  怪攻,怪命,怪防,怪暴率,怪暴伤,怪攻速, 怪火攻,怪冰攻,怪雷攻,怪毒攻, 怪火抗,怪冰抗,怪雷抗,怪毒抗]
test_data = [
    [100,1000,2000,20,150,30,10, 50,0,30,0, 20,0,10,0,0,
     50,800,20,10,100,1500, 0,0,0,0, 15,0,10,0],
    [250,800,3500,35,200,50,25, 80,60,0,0, 30,0,0,0,20,
     120,3000,40,15,150,1200, 40,0,0,0, 0,25,0,10],
    [500,600,8000,50,250,80,40, 0,150,0,100, 40,20,0,10,0,
     300,15000,60,20,200,1000, 0,0,80,0, 0,0,30,0],
    [80,1200,1500,10,100,15,0, 0,0,0,0, 10,0,0,0,0,
     40,600,10,5,100,1800, 0,0,0,0, 0,0,0,0],
    [1200,500,20000,60,300,120,60, 200,200,200,200, 60,30,30,30,30,
     800,50000,90,25,250,800, 200,200,200,200, 40,40,40,40],
]

START_ROW = 4

def write_formula_row(r):
    """写入第r行的伤害计算公式。返回引用列表。"""
    # 列变量（1-based）
    # 人物 A-P
    pA=1; pSpd=2; pHP=3; pCR=4; pCD=5; pDef=6; pPrc=7
    pFire=8; pIce=9; pThund=10; pPoison=11
    pRes=12; pRFire=13; pRIce=14; pRThund=15; pRPoison=16
    # 怪物 Q-AD → col 17-30
    mA=17; mHP=18; mDef=19; mCR=20; mCD=21; mSpd=22
    mFire=23; mIce=24; mThund=25; mPoison=26
    mRFire=27; mRIce=28; mRThund=29; mRPoison=30
    # 计算列 31-38
    cPNo=31; cPC=32; cPKn=33; cPKt=34
    cMNo=35; cMC=36; cMKn=37; cMKt=38

    def L(col): return f"{get_column_letter(col)}{r}"

    # ===== 人物攻击怪物 非暴击 =====
    # 物理基础 = 攻击力 * (1 - max(0, 怪防-穿透)/100)
    phys = f"{L(pA)}*(1-MAX(0,{L(mDef)}-{L(pPrc)})/100)"
    # 每种属性伤害 = 人物元素伤害 * (1 - 怪物对应抗性/100)
    fire  = f"{L(pFire)}*(1-{L(mRFire)}/100)"
    ice   = f"{L(pIce)}*(1-{L(mRIce)}/100)"
    thund = f"{L(pThund)}*(1-{L(mRThund)}/100)"
    poison= f"{L(pPoison)}*(1-{L(mRPoison)}/100)"
    p_non_crit = f"=MAX(1,ROUND({phys}+{fire}+{ice}+{thund}+{poison},0))"

    # 暴击 = 非暴击公式整体 × (1+暴伤/100)
    p_crit = f"=MAX(1,ROUND(({phys}+{fire}+{ice}+{thund}+{poison})*(1+{L(pCD)}/100),0))"

    # 击杀次数 = ceil(怪命 / 非暴击)
    p_kn = f"=IFERROR(CEILING({L(mHP)}/{get_column_letter(cPNo)}{r},1),\"\")"
    # 击杀时间 = 次数 * (1000/攻速) / 1000 = 次数 / 攻速
    p_kt = f"=IFERROR({get_column_letter(cPKn)}{r}*(1000/{L(pSpd)})/1000,\"\")"

    # ===== 怪物攻击人物 非暴击 =====
    # 通用减伤 = 1 / (1 + 通抗/100)  → 通抗100=50%, 通抗200=66.7%
    general = f"1/(1+{L(pRes)}/100)"
    # 物理基础 = 怪攻 × (1-玩家防/100) × 通用减伤
    m_phys = f"{L(mA)}*(1-{L(pDef)}/100)*{general}"
    # 每种怪元素伤害 = 怪元素攻击 × (1-玩家对应抗性/100) × 通用减伤
    m_fire  = f"{L(mFire)}*(1-{L(pRFire)}/100)*{general}"
    m_ice   = f"{L(mIce)}*(1-{L(pRIce)}/100)*{general}"
    m_thund = f"{L(mThund)}*(1-{L(pRThund)}/100)*{general}"
    m_poison= f"{L(mPoison)}*(1-{L(pRPoison)}/100)*{general}"
    m_non_crit = f"=MAX(1,ROUND({m_phys}+{m_fire}+{m_ice}+{m_thund}+{m_poison},0))"

    # 暴击
    m_crit = f"=MAX(1,ROUND(({m_phys}+{m_fire}+{m_ice}+{m_thund}+{m_poison})*(1+{L(mCD)}/100),0))"

    # 击杀次数 = ceil(玩家命 / 非暴击)
    m_kn = f"=IFERROR(CEILING({L(pHP)}/{get_column_letter(cMNo)}{r},1),\"\")"
    m_kt = f"=IFERROR({get_column_letter(cMKn)}{r}*(1000/{L(mSpd)})/1000,\"\")"

    formulas = [p_non_crit, p_crit, p_kn, p_kt, m_non_crit, m_crit, m_kn, m_kt]
    cols = [cPNo, cPC, cPKn, cPKt, cMNo, cMC, cMKn, cMKt]
    return list(zip(cols, formulas))


# 写入测试数据
for ridx, row in enumerate(test_data):
    r = START_ROW + ridx
    # 人物 A-P 16列 (index 0-15)
    for i in range(P_TOTAL):
        cell = ws.cell(row=r, column=i+1, value=row[i])
        cell.font = data_font; cell.alignment = center; cell.border = border
        cell.fill = fill_player
    # 怪物 Q-AD 14列 (index 16-29)
    for i in range(M_TOTAL - P_TOTAL):
        cell = ws.cell(row=r, column=P_TOTAL+1+i, value=row[P_TOTAL+i])
        cell.font = data_font; cell.alignment = center; cell.border = border
        cell.fill = fill_monster
    # 公式列
    for col, fml in write_formula_row(r):
        cell = ws.cell(row=r, column=col, value=fml)
        cell.font = data_font; cell.alignment = center; cell.border = border
        cell.fill = fill_calc
        if col in (33, 34, 37, 38):
            cell.number_format = "0.00"

# 预填 20 行空行（属性留空 + 公式预填）
EMPTY_ROWS = 20
for ridx in range(EMPTY_ROWS):
    r = START_ROW + len(test_data) + ridx
    for i in range(P_TOTAL):
        cell = ws.cell(row=r, column=i+1)
        cell.font = data_font; cell.alignment = center; cell.border = border
        cell.fill = fill_player
    for i in range(M_TOTAL - P_TOTAL):
        cell = ws.cell(row=r, column=P_TOTAL+1+i)
        cell.font = data_font; cell.alignment = center; cell.border = border
        cell.fill = fill_monster
    # 预填公式（IFERROR 包一层，空数据时显示空）
    for col, fml in write_formula_row(r):
        fml_safe = f"=IFERROR({fml[1:]},\"\")"  # 去掉开头的 =，外面重包 IFERROR
        # 上面 write_formula_row 已返回带 = 的完整公式；用 IFERROR 重包避免空行 DIV/0
        # 手动剥掉 = 再包
        core = fml[1:]
        fml_safe = f"=IFERROR({core},\"\")"
        cell = ws.cell(row=r, column=col, value=fml_safe)
        cell.font = data_font; cell.alignment = center; cell.border = border
        cell.fill = fill_calc
        if col in (33, 34, 37, 38):
            cell.number_format = "0.00"

# ===== 列宽 =====
widths = ([8,10,9,8,10,7,9, 8,8,8,8, 9,9,9,9,9]
          + [8,9,7,8,10,10, 8,8,8,8, 9,9,9,9]
          + [16,14,14,14,16,14,14,14])
for i, w in enumerate(widths, start=1):
    if i <= TOTAL_COLS:
        ws.column_dimensions[get_column_letter(i)].width = w

# 冻结窗格：前三行 + 前两部分列固定
freeze_at = f"{get_column_letter(M_TOTAL+1)}{START_ROW}"
ws.freeze_panes = freeze_at

# ===== 备注说明 =====
note_row = START_ROW + len(test_data) + EMPTY_ROWS + 2
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
c = ws.cell(row=note_row, column=1, value="公式说明（完整设计版）")
c.font = Font(name="微软雅黑", size=9, bold=True, color="5F6368")
notes = [
    "【人物攻击怪物】",
    "1. 物理基础 = 攻击力 × (1 - max(0, 怪物防御 - 物理穿透) / 100)",
    "2. 各属性实际 = 人物该属性伤害 × (1 - 怪物对应抗性 / 100)  （抗性超过100完全免疫该属性）",
    "3. 非暴击总伤害 = max(1, round(物理基础 + 火+冰+雷+毒属性实际, 0))",
    "4. 暴击总伤害 = max(1, round((物理基础 + 属性总和) × (1 + 暴击伤害/100), 0))",
    "5. 击杀次数 = CEILING(怪物最大生命 / 非暴击总伤害, 1)  （向上取整）",
    "6. 击杀时间(秒) = 击杀次数 × (1000 / 攻击速度) / 1000",
    "",
    "【怪物攻击人物】",
    "7. 通用减伤系数 = 1 / (1 + 通用抗性/100)  （抗性100=50%、抗性200=66.7%，平滑递减免伤）",
    "8. 物理基础 = 怪物攻击 × (1 - 玩家防御 / 100) × 通用减伤系数",
    "9. 怪物属性实际 = 该元素攻击 × (1 - 玩家对应抗性 / 100) × 通用减伤系数",
    "10. 非暴击总伤害 = max(1, round(物理基础 + 火+冰+雷+毒属性实际, 0))",
    "11. 暴击总伤害 = max(1, round(总和 × (1 + 怪物暴击伤害/100), 0))",
    "12. 击杀次数 = CEILING(玩家最大生命 / 非暴击总伤害, 1)",
    "13. 击杀时间(秒) = 击杀次数 × (1000 / 怪物攻击速度) / 1000",
    "",
    "【填写提示】",
    "14. 暴击率/暴击伤害/防御/穿透/抗性/属性伤害 均按百分比数值填写（填20表示20%，填100表示100%）",
    "15. 第三部分为 Excel 公式自动计算，修改第一、二部分数据即会重算",
]
for i, n in enumerate(notes, start=1):
    ws.merge_cells(start_row=note_row+i, start_column=1, end_row=note_row+i, end_column=30)
    cell = ws.cell(row=note_row+i, column=1, value=n)
    cell.font = Font(name="微软雅黑", size=9, color="5F6368")

out_path = r"D:\ShotsGame\tools\战斗伤害计算表.xlsx"
wb.save(out_path)
print(f"已生成: {out_path}")

# ===== Python 复算，便于用户核对 =====
print("\n===== 5条测试数据校验（Python复算，Excel应完全一致）=====")
header = f"{'行':<3} {'人→怪非暴':<10} {'人→怪暴':<10} {'人杀次':<7} {'人杀秒':<7} {'怪→人非暴':<10} {'怪→人暴':<10} {'怪杀次':<7} {'怪杀秒':<7}"
print(header)
for ridx, row in enumerate(test_data, start=START_ROW):
    (pA,pSpd,pHP,pCR,pCD,pDef,pPrc,
     pFire,pIce,pThund,pPoison,
     pRes,pRFire,pRIce,pRThund,pRPoison,
     mA,mHP,mDef,mCR,mCD,mSpd,
     mFire,mIce,mThund,mPoison,
     mRFire,mRIce,mRThund,mRPoison) = row

    # 人物攻击怪物
    phys = pA * (1 - max(0, mDef-pPrc)/100)
    f  = pFire  * (1 - mRFire/100)
    ic = pIce   * (1 - mRIce/100)
    th = pThund * (1 - mRThund/100)
    po = pPoison * (1 - mRPoison/100)
    total = phys + f + ic + th + po
    p_non = max(1, round(total))
    p_cri = max(1, round(total * (1 + pCD/100)))
    p_kn  = math.ceil(mHP / p_non)
    p_kt  = p_kn * (1000/pSpd) / 1000

    # 怪物攻击人物
    gen = 1 / (1 + pRes/100)
    m_phys = mA * (1 - pDef/100) * gen
    mf  = mFire  * (1 - pRFire/100) * gen
    mi  = mIce   * (1 - pRIce/100) * gen
    mt  = mThund * (1 - pRThund/100) * gen
    mp  = mPoison * (1 - pRPoison/100) * gen
    m_total = m_phys + mf + mi + mt + mp
    m_non = max(1, round(m_total))
    m_cri = max(1, round(m_total * (1 + mCD/100)))
    m_kn  = math.ceil(pHP / m_non)
    m_kt  = m_kn * (1000/mSpd) / 1000

    print(f"{ridx:<3} {p_non:<10} {p_cri:<10} {p_kn:<7} {p_kt:<7.2f} {m_non:<10} {m_cri:<10} {m_kn:<7} {m_kt:<7.2f}")
